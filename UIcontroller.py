import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QMainWindow
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QDesktopServices
from UI import Ui_MainWindow
import openpyxl
import os
import subprocess
import numpy as np
import pandas as pd
import csv
from openpyxl import load_workbook




class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)       
        self.setup_control()
        self.file_name = ''
        self.file_name2 = ''
        
        
    def setup_control(self):
        self.ui.Open.clicked.connect(self.buttonClicked)
        self.ui.Open2.clicked.connect(self.buttonClicked2)
        self.ui.Cancel.clicked.connect(self.Cancel)
        self.ui.OK.clicked.connect(self.OK)
        
        
    def buttonClicked(self):
        # file_name, _ = QFileDialog.getOpenFileName(self,'選取檔案')
        # workbook1 = openpyxl.load_workbook(file_name) 
        self.file_name, _ = QFileDialog.getOpenFileName(self, "選取檔案", "", "TEXT Files (*.csv)") 
        self.ui.lineEdit.setText(self.file_name)
        print(self.file_name)
        
        
    def buttonClicked2(self):
        # file_name, _ = QFileDialog.getOpenFileName(self,'選取檔案')
        # workbook2 = openpyxl.load_workbook(file_name) 
        self.file_name2, _ = QFileDialog.getOpenFileName(self, "選取檔案", "", "TEXT Files (*.txt)") 
        self.ui.lineEdit_2.setText(self.file_name2)
        print(self.file_name2)
       
        
    def Cancel(self):
        self.close()
        
        
    def OK(self):
        # CSV檔案轉成TXT文字檔
        # TXT檔案名稱
        txt_file_name = 'sample.txt'
        # 讀取CSV檔案並寫入TXT檔案
        with open(self.file_name, mode='r', encoding='utf-8') as aaa:
            csv_reader = csv.reader(aaa)
            # print(csv_reader)
            with open(txt_file_name, mode='w', encoding='utf-8') as txt_file:
                for row in csv_reader:
                    txt_file.write(' '.join(row) + '\n')
                # print(row)

        # 指定要刪除的行號
        lines_to_delete = {1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19}    # 行號從1開始計數 ,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19

        # 讀取檔案
        with open('sample.txt', 'r', encoding='utf-8') as file:
            lines = file.readlines()
            # print(lines)
            
        # 重新寫入檔案，除了指定要刪除的行
        with open('change.txt', 'w', encoding='utf-8') as file:
            for index, line in enumerate(lines, start=1):
                if index not in lines_to_delete:
                    file.write(line)   
                    
                    
                    
                    
                    
                                    
        # 讀取檔案                
        with open('change.txt', 'r', encoding='utf-8') as file:
            rread = file.readlines()
            #刪除每行第x列
            new_lines = []
            for line in rread:
                parts = line.split()
                new_parts = [part for i, part in enumerate(parts) if i in [1]]
                new_lines.append(' '.join(new_parts))               
                
        with open('DBM.txt', 'w', encoding='utf-8') as file:
            for line in new_lines:    
                file.write(line + '\n')
                
        # 讀取txt檔案
        df = pd.read_csv('DBM.txt', header=None, dtype=str)  # 假設txt檔案沒有標題行

        # 將文字轉換為數值，這裡假設所有數據都可以轉換
        df = df.apply(pd.to_numeric, errors='coerce')

        # 載入現有的Excel工作簿
        wb = load_workbook('turn.xlsx')
        ws = wb['工作表1']  # 替換成你的工作表名稱

        
        for index, row in df.iterrows():
            ws.cell(row=index+1, column=1, value=str(row[0]))  # 寫入第一列

       
        wb.save('turn.xlsx')                
                 
                
        with open(self.file_name2, 'r',  encoding='utf-8') as bbb:            
            rread = bbb.readlines()
            #刪除每行第x列
            new_lines = []
            for line in rread:
                parts = line.split()
                new_parts = [part for i, part in enumerate(parts) if i in [14]]
                new_lines.append(' '.join(new_parts))                    
                        
        with open('S1.txt', 'w', encoding='utf-8') as file:
            for line in new_lines:    
                file.write(line + '\n')
                
        # 讀取txt檔案
        df = pd.read_csv('S1.txt', header=None, dtype=str)  #txt檔案沒有標題行

        # 將文字轉換為數值，這裡假設所有數據都可以轉換
        df = df.apply(pd.to_numeric, errors='coerce')

        # 載入現有的Excel工作簿
        wb = load_workbook('turn.xlsx')
        ws = wb['工作表1']  # 替換成你的工作表名稱

        # 從第二行開始寫入數據
        for index, row in df.iterrows():
            ws.cell(row=index+1, column=5, value=str(row[0]))  # 假設你想寫入第五列

        # 儲存工作簿
        wb.save('turn.xlsx')                 
                
                
                
                
                
                
                
                             
        with open(self.file_name2, 'r',  encoding='utf-8') as bbb:
            rread = bbb.readlines()
            #刪除每行第x列
            new_lines = []
            for line in rread:
                parts = line.split()
                new_parts = [part for i, part in enumerate(parts) if i in [17]]
                new_lines.append(' '.join(new_parts))         
                        
        with open('S2.txt', 'w', encoding='utf-8') as file:
            for line in new_lines:    
                file.write(line + '\n') 

        # 讀取txt檔案
        df = pd.read_csv('S2.txt', header=None, dtype=str)  # 假設txt檔案沒有標題行

        # 將文字轉換為數值，這裡假設所有數據都可以轉換
        df = df.apply(pd.to_numeric, errors='coerce')

        # 載入現有的Excel工作簿
        wb = load_workbook('turn.xlsx')
        ws = wb['工作表1']  # 替換成你的工作表名稱

        # 從第二行開始寫入數據
        for index, row in df.iterrows():
            ws.cell(row=index+1, column=6, value=str(row[0]))  # 假設你想寫入第六列

        # 儲存工作簿
        wb.save('turn.xlsx')                 
                






# 創建應用實例
app = QApplication([])

# 創建窗口實例
demo = MainWindow()
demo.show()

# 運行應用
app.exec_()